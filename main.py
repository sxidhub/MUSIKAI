from pytubefix import YouTube, exceptions # Import exceptions for better error handling
import vlc
import tkinter as tk
from tkinter import ttk # Import ttk for themed scrollbar
import openpyxl
import os
import requests # para descargar la imagen
from PIL import Image, ImageTk#poner imagenes en el tkinter
from pythonds3 import Queue
import ast #esta es para convertir las playlist de exel en listas
import random as r
import time as t
import shutil

#Globales ----------------
ventana_principal = tk.Tk()
ventana_principal.title("MUZIKAI")
ventana_principal.geometry("1200x800")
ventana_principal.minsize(600,400)
ventana_principal.config(bg = "#014942")
ruta_main=""
usuario_activo = ""
archivo_audio_path_temp = ""
global current_vlc_player # Declara una nueva variable global para el reproductor
current_vlc_player = None # Inicialízala a None
global progress_scale
progress_scale = None
global update_progress_job
update_progress_job = None
global imagen_label
ventana_descarga = None
admin = ""
premium = None
cola_reproduccion = Queue()
playlists_registro=[]

#Crear Archivos iniciales ----------------
try:
    archivo=openpyxl.load_workbook('canciones.xlsx')
    archivo.save("canciones.xlsx")
    hoja = archivo.active
except FileNotFoundError:
    archivo=openpyxl.Workbook()
    hoja = archivo.active
    hoja.title = "USUARIOS"
    archivo.save("canciones.xlsx")

def agregar_cola(archivo_audio):
    cola_reproduccion.enqueue(archivo_audio)
    print(archivo_audio)
    etiqueta = tk.Label (ventana_principal,
                            text="Canción añadida a la cola",
                            foreground="#008B7E",
                            bg="#014942",
                            font="Corbel 18",
                            )
    etiqueta.pack(pady=2)
    ventana_principal.after(1500, lambda: destruir_etiqueta_si_existe(etiqueta))

def verificar_cola(etiqueta_cancion_actual): #Tiene Color
    if current_vlc_player:
        if current_vlc_player.get_state() == vlc.State.Ended:
            if cola_reproduccion.size()!=0:
                administrar_audio(cola_reproduccion.dequeue(),etiqueta_cancion_actual)
    ventana_principal.after(480, lambda: verificar_cola(etiqueta_cancion_actual)) 

def reproducir_siguiente(etiqueta_cancion_actual):#Tiene Color
    if cola_reproduccion.size()!=0:
        administrar_audio(cola_reproduccion.dequeue(),etiqueta_cancion_actual)
    else:
        etiqueta = tk.Label (ventana_principal,
                        text="No hay canción siguiente",
                        foreground="#FF0000",
                        bg="#014942",
                        font="Corbel 18",
                        )
        etiqueta.pack(pady=2)
        ventana_principal.after(1500, lambda: destruir_etiqueta_si_existe(etiqueta))

def eliminar_cancion(canciones_registro, archivo_audio,frame_canciones_placeholder,etiqueta_cancion_actual,frame_playlists_placeholder,frame_playlist_canciones_placeholder):
    global hoja
    global imagen_label
    global current_vlc_player
    hoja = archivo[usuario_activo]
    nombre_eliminar = os.path.basename(archivo_audio)

    # Limpiar la hoja si ya tiene datos
    i=1
    while(hoja["A"+str(i)].value is not None):
        hoja["A" + str(i)].value = None
        hoja["B" + str(i)].value = None
        i = i+1

    i = 1
    # Escribir solo las canciones que no se van a eliminar
    for link, nombre in canciones_registro.items():
        if nombre != nombre_eliminar:
            hoja["B" + str(i)].value = nombre
            hoja["A" + str(i)].value = link
            i += 1
    archivo.save("canciones.xlsx") #GUARDA CAMBIOS HECHOS EN EXCEL
    #Eliminar cancion

    print("main:",playlists_registro)
    print(nombre_eliminar)
    for playlist in playlists_registro:
        print(playlist)
        eliminar_cancion_playlist(playlist,nombre_eliminar)
    mostrar_playlist_canciones(playlist,frame_playlist_canciones_placeholder,etiqueta_cancion_actual)

    ruta_archivo = os.path.join(ruta_main, nombre_eliminar)
    try:
        if os.path.exists(ruta_archivo):
            os.remove(ruta_archivo)
            ruta_archivo = ruta_archivo.replace('.mp3', '.jpg')
            os.remove(ruta_archivo)
            print(f"Se eliminó la canción '{nombre_eliminar}' con éxito.")
        else:
            print(f"Archivo no encontrado: {ruta_archivo}")
    except Exception as e:
            current_vlc_player.stop()
            current_vlc_player.release()
            current_vlc_player = None
            if os.path.exists(ruta_archivo):
                os.remove(ruta_archivo)
                ruta_archivo = ruta_archivo.replace('.mp3', '.jpg')
                os.remove(ruta_archivo)
                print(f"Se eliminó la canción '{nombre_eliminar}' con éxito.")
            else:
                print(f"Archivo no encontrado: {ruta_archivo}")

    key_borrar = None
    for keys,values in canciones_registro.items():
        if nombre_eliminar == values:
            key_borrar = keys
    del(canciones_registro[key_borrar])
    imagen_label.config(image='') #Borrar la imagen
    etiqueta_cancion_actual.config(text="") #Borrar el nombre
    mostrar_canciones(canciones_registro, frame_canciones_placeholder,etiqueta_cancion_actual,frame_playlists_placeholder,frame_playlist_canciones_placeholder)

def eliminar_playlist(playlist_eliminar,frame_playlists_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder):
    global hoja
    global imagen_label
    global current_vlc_player
    global archivo
    global playlists_registro
    hoja = archivo[usuario_activo]

    # Limpiar la hoja si ya tiene datos

    while(hoja["C"+str(j)].value is not None):
        hoja["C" + str(j)].value = None

    archivo.save("canciones.xlsx") #GUARDA CAMBIOS HECHOS EN EXCEL

    # Escribir solo las canciones que no se van a eliminar
    j=1
    for lista in playlists_registro:
        if lista == playlist_eliminar:
            j=j-1
        else:
            hoja["C" + str(j)].value = str(lista)
        j=j+1

    archivo.save("canciones.xlsx") #GUARDA CAMBIOS HECHOS EN EXCEL
    print("LISTA ELIMINADA")
    
    mostrar_playlists(frame_playlists_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder)

def importar_canciones(canciones_registro):
    global hoja
    hoja = archivo[usuario_activo]

    if hoja["A1"].value is not None:
        i=1
        while(hoja["A"+str(i)].value is not None):
            link = hoja["A" + str(i)].value
            nombre = hoja["B" + str(i)].value
            canciones_registro[link] = nombre
            i = i+1

    return canciones_registro

def importar_playlists():
    global playlists_registro
    playlists_registro=[]
    global hoja
    hoja = archivo[usuario_activo]
    lista_aux = []
    ultima_fila_columna_c = 0
    
    if hoja["C1"].value is not None:
        i=1
        ultima_fila_columna_c = 0
        while(hoja["C"+str(i)].value is not None):
            ultima_fila_columna_c = i
            i = i+1
        for j in range(1,ultima_fila_columna_c+1):
            lista_aux = ast.literal_eval(hoja["C"+str(j)].value)
            playlists_registro.append(lista_aux)


def limpiar_frame(frame):
    for widget in frame.winfo_children():
        widget.destroy()


def verificar_biblioteca(entrada,lista):
    global linea_clave
    global hoja
    hoja = archivo[usuario_activo]
    lista = {}
    importar_canciones(lista)
    
    Existe=False
    i=1
    if hoja["A1"].value is not None:
        for key in lista:
            #importar_canciones[i-1].strip()==url.strip() or -1 !=(url.strip()).find(importar_canciones[i-1].strip()) or -1 !=(importar_canciones[i-1].strip()).find(url.strip()):
            #Verificaciones extra del otro code para que no tome el mismo link si le añades el cacho de cuando viene de una playlist en youtube
            if key == entrada or -1!=entrada.find(key) or -1!=key.find(entrada):
                Existe = True
                print("Tu cancion ya esta en la biblioteca") # Corrected typo
                linea_clave=i
            i=i+1

    return Existe


def menu_agregar_playlist(ventana_p,frame_playlists_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder):#Tiene Color
    global ventana_descarga
    if ventana_descarga is not None and ventana_descarga.winfo_exists():
        ventana_descarga.focus()
        return
    
    ventana_secundaria = tk.Toplevel(ventana_p)
    ventana_descarga = ventana_secundaria
    ventana_secundaria.title("MUZIKAI")
    ventana_secundaria.geometry("400x200")
    ventana_secundaria.resizable(False,False)
    ventana_secundaria.config(bg = "#02695F")
    ventana_descarga.protocol("WM_DELETE_WINDOW", lambda: cerrar_ventana_descarga())

    etiqueta1 = tk.Label (ventana_secundaria,
                          text="Nombre de la playlist:",
                          foreground="#FFFFFF",
                          bg="#02695F",
                          width="19",
                          height="1",
                          font="Corbel 25",
                          relief="groove")

    entrada1 = tk.Entry (ventana_secundaria,
                          foreground="#FFFFFF",
                          bg="#02695F",
                          width="30",
                          font="Corbel 20",
    )

    boton_agregar = tk.Button(ventana_secundaria,
                          text="Agregar",
                          bg="#02695F",
                          foreground="#FFFFFF",
                          font="Corbel 10",
                          command=lambda: agregar_playlist(entrada1,ventana_p,ventana_secundaria,frame_playlists_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder)
    )
    etiqueta1.pack(pady=10)
    entrada1.pack(pady=10)
    boton_agregar.pack()

def menu_descarga(entrada,lista,ventana_p,frame_canciones_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder,frame_playlists_placeholder): #Tiene Color
    global ventana_descarga

    if ventana_descarga is not None and ventana_descarga.winfo_exists():
        ventana_descarga.focus()
        return
    ventana_secundaria = tk.Toplevel(ventana_p)
    ventana_descarga = ventana_secundaria
    ventana_secundaria.title("MUZIKAI")
    ventana_secundaria.geometry("400x200")
    ventana_secundaria.resizable(False,False)
    ventana_secundaria.config(bg = "#02695F")
    ventana_descarga.protocol("WM_DELETE_WINDOW", lambda: cerrar_ventana_descarga())
    checkbox_estado = tk.BooleanVar()

    etiqueta1 = tk.Label (ventana_secundaria,
                          text="Nombre de la cancion:",
                          foreground="#FFFFFF",
                          bg="#02695F",
                          width="19",
                          height="1",
                          font="Corbel 25",
                          relief="groove"
    )

    entrada1 = tk.Entry (ventana_secundaria,
                          foreground="#FFFFFF",
                          bg="#02695F",
                          width="30",
                          font="Corbel 20",
    )

    checkbox1 = tk.Checkbutton(ventana_secundaria,
                            text="Descargar",
                            foreground="#000000",
                            bg="#02695F",
                            width="30",
                            font="Corbel 15",
                            variable=checkbox_estado,
                            onvalue=True,
                            offvalue=False,
    )

    boton_ver = tk.Button(ventana_secundaria,
                          text="Ingresar",
                          bg="#02695F",
                          activebackground="#149E9E",
                          foreground="#FFFFFF",
                          font="Corbel 10",
                          command=lambda: cancion(entrada,lista,ventana_secundaria,ventana_p,entrada1,checkbox_estado,frame_canciones_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder,frame_playlists_placeholder) # Use placeholder
    )

    etiqueta1.pack()
    entrada1.pack()
    checkbox1.pack()
    boton_ver.pack()

def cerrar_ventana_descarga():
    global ventana_descarga
    if ventana_descarga is not None:
        ventana_descarga.destroy()
        ventana_descarga = None




def menu_usuarios():#Tiene Color
    global ventana_descarga
    if ventana_descarga is not None and ventana_descarga.winfo_exists():
        ventana_descarga.focus()
        return
    global hoja
    hoja = archivo["USUARIOS"]

    ventana_secundaria = tk.Toplevel(ventana_principal)
    ventana_descarga = ventana_secundaria
    ventana_secundaria.title("MUZIKAI - Gestión de Usuarios")
    ventana_secundaria.geometry("700x450")
    ventana_secundaria.config(bg="#02695F")
    ventana_descarga.protocol("WM_DELETE_WINDOW", lambda: cerrar_ventana_descarga())

    ventana_secundaria.grid_rowconfigure(0, weight=1)
    ventana_secundaria.grid_columnconfigure(0, weight=1)

    frame_usuarios = tk.Frame(ventana_secundaria, bg="#9312DD")
    frame_usuarios.grid(row=0, column=0, sticky="nsew")

    frame_usuarios.grid_rowconfigure(0, weight=0)
    frame_usuarios.grid_rowconfigure(1, weight=1)
    frame_usuarios.grid_columnconfigure(0, weight=1)

    label_usuarios = tk.Label(frame_usuarios,
                               text="Menú de Usuarios",
                               foreground="#FFFFFF",
                               bg="#9312DD",
                               justify="center",
                               font="Corbel 20 bold",
                               pady=10)
    label_usuarios.grid(row=0, column=0, sticky="ew", padx=10, pady=5)

    frame_treeview_container = tk.Frame(frame_usuarios, bg="#DD12B1")
    frame_treeview_container.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)

    frame_treeview_container.grid_rowconfigure(0, weight=1)
    frame_treeview_container.grid_columnconfigure(0, weight=1)
    frame_treeview_container.grid_columnconfigure(1, weight=0)

    style = ttk.Style()
    style.theme_use("clam")

    style.configure("Treeview",
                    background="#E1E1E1",
                    foreground="black",
                    rowheight=25,
                    fieldbackground="#E1E1E1")
    style.map("Treeview",
              background=[('selected', '#347083')],
              foreground=[('selected', 'white')])

    style.configure("Treeview.Heading",
                    font=("Corbel", 12, "bold"),
                    background="#5A0A8C",
                    foreground="white",
                    relief="flat")
    style.map("Treeview.Heading", background=[('active', '#6A1A9C')])

    columns = ("usuario", "contrasenia", "permisos", "accion")
    tree = ttk.Treeview(frame_treeview_container, columns=columns, show="headings", selectmode="browse")

    tree.heading("usuario", text="Usuario", anchor="w")
    tree.heading("contrasenia", text="Contraseña", anchor="w")
    tree.heading("permisos", text="Permisos", anchor="w")
    tree.heading("accion", text="Borrar", anchor="center")

    tree.column("usuario", width=160, minwidth=100, stretch=tk.YES, anchor="w")
    tree.column("contrasenia", width=160, minwidth=100, stretch=tk.YES, anchor="w")
    tree.column("permisos", width=90, minwidth=50, stretch=tk.YES, anchor="w")
    tree.column("accion", width=60, minwidth=50, stretch=tk.NO, anchor="center")

    vsb = ttk.Scrollbar(frame_treeview_container, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)

    vsb.grid(row=0, column=1, sticky="ns")
    tree.grid(row=0, column=0, sticky="nsew")

    def cargar_usuarios_en_treeview():
        for item in tree.get_children():
            tree.delete(item)

        i = 1
        while True:
            usuario_val = hoja["A" + str(i)].value
            contrasenia_val = hoja["B" + str(i)].value
            permisos_val = hoja["C" + str(i)].value

            if usuario_val is None:
                break

            if i == 1 and str(usuario_val).strip().lower() == 'usuario':
                i += 1
                continue

            tree.insert("", "end", values=(usuario_val, contrasenia_val, permisos_val, "X"))
            i += 1

    cargar_usuarios_en_treeview()

    def on_tree_click(event):
        item_id = tree.identify_row(event.y)
        column_id = tree.identify_column(event.x)

        if item_id and column_id == "#4":  # Column index for "accion" (Borrar)
            values = tree.item(item_id, 'values')
            if values:
                usuario_a_eliminar = values[0]  # User is in the first position
                permisos_del_usuario_seleccionado = values[2] # Permissions are in the third position

                if permisos_del_usuario_seleccionado.lower() == "usuario":
                    eliminar_sesion(usuario_a_eliminar)
                    cargar_usuarios_en_treeview()
                else:
                    mensaje_error(ventana_principal,"No puedes borrar administradores")


    tree.bind("<Button-1>", on_tree_click)

def menu_canciones_playlist(cancion_nombre_mp3, etiqueta_cancion_actual, frame_playlists_placeholder_principal,frame_playlist_canciones_placeholder):#Tiene Color
    global ventana_descarga
    if ventana_descarga is not None and ventana_descarga.winfo_exists():
        ventana_descarga.focus()
        return
    global playlists_registro
    ventana_secundaria = tk.Toplevel(ventana_principal)
    ventana_descarga = ventana_secundaria
    ventana_secundaria.title("MUZIKAI")
    ventana_secundaria.geometry("250x300")
    ventana_secundaria.resizable(False,False)
    ventana_secundaria.config(bg = "#02695F")
    ventana_descarga.protocol("WM_DELETE_WINDOW", lambda: cerrar_ventana_descarga())
    

    frame_playlists = tk.Frame(ventana_secundaria, #Frame para las playlists
                        bg="#00776B",
                        width=250,
                        height=300
                        )
    
    # --- Copia del scrolling pero para frame_playlist ---
    canvas2 = tk.Canvas(frame_playlists,  #Canva para contener canciones dentro de frame_canciones_contenedor
                            bg="#00776B", 
                            width=235,
                            highlightthickness=0
                            ) 
    
    label_playlist = tk.Label(frame_playlists,
                              text= "A que playlist agregar",
                              foreground="#FFFFFF",
                              bg="#00776B",
                              justify="center",
                              font="Corbel 18",
                              )
    
    scrollbar2 = ttk.Scrollbar(frame_playlists,  #Scrollbar dentro de frame_canciones_contenedor vinculado al canvas
                            orient="vertical", 
                            command=canvas2.yview
                            )
    # Create an inner frame within the Canvas to hold the song entries, and this is where your grid layout for songs will go.
    frame_playlists_placeholder = tk.Frame(canvas2, 
                            width=200,
                            bg="#00776B"
                            ) 

    # Add the inner frame to a window in the canvas
    canvas2_window_id = canvas2.create_window((0, 0), 
                            window=frame_playlists_placeholder,
                            width=200, 
                            anchor="nw")

    # Function to adjust the width of the frame_canciones_placeholder window within the canvas and to update the scrollregion.
    def on_canvas_configure2(event):
        canvas2.itemconfig(canvas2_window_id, width=event.width) # Update the width of the window item to match the canvas width
        canvas2.configure(scrollregion=canvas2.bbox("all"))# Update the scrollregion. This is critical for scrolling to work.

    
    canvas2.bind('<Configure>', on_canvas_configure2) # Bind the canvas's <Configure> event to this function

    # Bind the inner frame's <Configure> event to update the scrollregion. This is important when content within `frame_canciones_placeholder` changes size.
    frame_playlists_placeholder.bind("<Configure>", lambda o: canvas2.configure(scrollregion=canvas2.bbox("all")))

    # --- End of Scrolling Implementation ---

    canvas2.configure(yscrollcommand=scrollbar2.set)
    frame_playlists.pack_propagate(False)

    label_playlist.pack(pady=5)
    canvas2.pack(side="left", fill="both", expand=True)
    scrollbar2.pack(side="right", fill="y")
    frame_playlists.pack()

    limpiar_frame(frame_playlists_placeholder)
    frame_playlists_placeholder.grid_columnconfigure(0, weight=1) 
    frame_playlists_placeholder.grid_columnconfigure(1, weight=1) 

    importar_playlists()

    i=0
    for playlist in playlists_registro:
        playlist_nombre = playlist[0]

        song_row_frame = tk.Frame(frame_playlists_placeholder, bg="#008B7E") # Greenish background for song rows
        song_row_frame.grid(row=i, column=0, columnspan=2, sticky="ew", padx=5, pady=3) # Span all 2
        
        #Solo van a ser 2
        song_row_frame.grid_columnconfigure(0, weight=0) # Para el boton agregar
        song_row_frame.grid_columnconfigure(1, weight=1) # Para el nombre de la canción

        
        boton_agregar_playlist = tk.Button(song_row_frame, text="+", foreground="#FFFFFF", bg="#008B7E",
                                    font="Corbel 10", justify="center", activebackground="#0273A7",
                                    activeforeground="#FFFFFF", 
                                    command=lambda a=playlist, b=cancion_nombre_mp3: agregar_cancion_playlist(a, b, ventana_secundaria, frame_playlists_placeholder_principal, etiqueta_cancion_actual,frame_playlist_canciones_placeholder),
                                    width=2) # Fixed width for consistent button size
        
        etiqueta_playlist = tk.Label(song_row_frame, text=playlist_nombre, foreground="#FFFFFF", bg="#008B7E",
                                    justify="left", font="Corbel 16", anchor="w", wraplength=400) # Added wraplength

        boton_agregar_playlist.grid(row=0, column=0, padx=(5,2), pady=2, sticky="w")
        etiqueta_playlist.grid(row=0, column=1, padx=2, pady=2, sticky="ew")
        i=i+1

    frame_playlists_placeholder.update_idletasks() # Ensure widgets are drawn and sizes calculated
    frame_playlists_placeholder.master.config(scrollregion=frame_playlists_placeholder.master.bbox("all"))

def agregar_cancion_playlist(playlist,cancion_nombre_mp3,ventana_secundaria,frame_playlists_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder):
    global hoja
    global archivo
    lista_aux = []
    hoja = archivo[usuario_activo]
    ventana_principal.after(1000, lambda: cerrar_ventana(ventana_secundaria))
    if hoja["C1"].value is not None:
        i=1
        while(hoja["C"+str(i)].value is not None):
            lista_aux = ast.literal_eval(hoja["C"+str(i)].value)
            if playlist[0] == lista_aux[0]:
                if cancion_nombre_mp3 in lista_aux:
                    mensaje_error(ventana_principal,"Canción ya en playlist")
                    return
                lista_aux.append(cancion_nombre_mp3)
                hoja["C"+str(i)].value = str(lista_aux)
                mensaje_error(ventana_principal,"Canción agregada")
            i = i+1
        archivo.save("canciones.xlsx")
        importar_playlists()
        mostrar_playlists(frame_playlists_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder)

def agregar_playlist(entrada1,ventana_p,ventana_s,frame_playlists_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder):
    global hoja
    global playlists_registro
    hoja = archivo[usuario_activo]
    ventana_p.after(1000, lambda: cerrar_ventana(ventana_s))
    nombre_playlist = entrada1.get()

    if not nombre_playlist.strip():
        print("ERROR: El nombre no puede estar vacío")
        mensaje_error(ventana_s, "Nombre vacío")
        return
    if len(nombre_playlist) > 20:
        print("ERROR: Nombre muy largo.")
        mensaje_error(ventana_s, "Nombre muy largo")
        return
    
    for playlist in playlists_registro:
        if playlist[0] == nombre_playlist:
            print("No se pueden repetir nombres")
            mensaje_error(ventana_s,"No repetir nombre")
            return
    #APARTIR DE AQUI LE MOVI
    playlist = [nombre_playlist]
    #Estas lineas guardan la playlist en la columna C del usuario
    ultima_fila_columna_c = 0
    i=1
    while(hoja["C"+str(i)].value is not None):
        ultima_fila_columna_c = i
        i= i+1
    hoja["C"+str(ultima_fila_columna_c+1)].value=str(playlist)
    archivo.save("canciones.xlsx")
    
    importar_playlists()
    mostrar_playlists(frame_playlists_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder)


def cancion(entrada,lista,ventana_secundaria,ventana_p,entrada1,checkbox,frame_canciones_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder,frame_playlists_placeholder): # Use placeholder
    global hoja
    global current_vlc_player
    global archivo_audio_path_temp
    global imagen_label
    hoja = archivo[usuario_activo]
    
    try:
        current_vlc_player.stop()
        current_vlc_player.release()
        current_vlc_player = None
        imagen_label.config(image='') # Quita la imagen
        imagen_label.image = None       
        os.remove(archivo_audio_path_temp)
        archivo_audio_path_temp = archivo_audio_path_temp.replace('.mp3', '.jpg')
        os.remove(archivo_audio_path_temp)
        print(f"DEBUG: Archivo '{archivo_audio_path_temp}' eliminado exitosamente.")
    except PermissionError as e:
        print(f"ERROR: No se pudo eliminar el archivo '{archivo_audio_path_temp}'. Aún en uso: {e}")
    except Exception as e:
        print(f"ERROR inesperado al eliminar '{archivo_audio_path_temp}': {e}")


    ventana_p.after(1000, lambda: cerrar_ventana(ventana_secundaria))
    link_cancion = entrada.get()
    nombre_cancion = entrada1.get() # Get value from Entry widget

    if not link_cancion.strip():
        print("ERROR: El campo del link de YouTube está vacío.")
        mensaje_error(ventana_secundaria, "Espacio vacio")
        return
    
    if len(nombre_cancion)>45:
        print("ERROR: Nombre muy largo.")
        mensaje_error(ventana_secundaria, "Nombre muy largo")
        return
    
    yt = None
    try:
        yt = YouTube(link_cancion)
        if yt.age_restricted:
            print("Este video tiene restricción de edad, no se puede procesar.")
            validador = False
            return
    except exceptions.RegexMatchError:
        print(f"ERROR: El link '{link_cancion}' no es un link de YouTube válido.")
        mensaje_error(ventana_secundaria, "Link de YouTube inválido. Intente con otro.")
        return
    except Exception as e:
        print(f"ERROR: No se pudo procesar el link de YouTube '{link_cancion}': {e}")
        mensaje_error(ventana_secundaria, f"Error al procesar el link: {e}")
        return 

    if checkbox.get() == False: # Solo escuchar (reproducción temporal)
        try:
            #INICIO CACHO QUE DESCARGA IMAGENES
            thumbnail_url = yt.thumbnail_url
            response = requests.get(thumbnail_url)
            if response.status_code == 200:
                with open(ruta_main+'/'+'.jpg', "wb") as f:
                    f.write(response.content)
                print(f"Miniatura guardada en: {ruta_main}")
            else:
                print("Error al descargar la miniatura.")
            #FINAL CACHO QUE DESCARGA IMAGENES
            audio = yt.streams.filter(only_audio=True).first()
            if audio is None:
                print("ERROR: No se encontró stream de audio para el video.")
                mensaje_error(ventana_secundaria, "No se encontró audio para este video.")
                return

            archivo_audio_path = audio.download(output_path=ruta_main, filename=".mp3")
            
            administrar_audio(archivo_audio_path,etiqueta_cancion_actual)

            archivo_audio_path_temp = archivo_audio_path
            # Ya no se pasa la duración, se monitorea el estado
            solo_escuchar(archivo_audio_path,nombre_cancion,etiqueta_cancion_actual) 
        except Exception as e:
            print(f"ERROR durante descarga/reproducción temporal: {e}")
            mensaje_error(ventana_secundaria, f"Error al procesar la canción: {e}")

    else: 
        if nombre_cancion == "":
            print("Ingrese un nombre")
            mensaje_error(ventana_secundaria, "Ingrese un nombre")
            return    

        En_biblioteca = verificar_biblioteca(link_cancion,lista)
        
        validador = True
        if En_biblioteca is False:
            i=1
            ultima_fila_columna_a = 0
            while(hoja["A"+str(i)].value is not None):
                ultima_fila_columna_a = i
                i = i+1
            hoja["A"+str(ultima_fila_columna_a+1)].value=link_cancion
            hoja["B"+str(ultima_fila_columna_a+1)].value=nombre_cancion+'.mp3' # Use nombre_cancion
            archivo.save("canciones.xlsx")
            
            #INICIO CACHO QUE DESCARGA IMAGENES
            thumbnail_url = yt.thumbnail_url
            response = requests.get(thumbnail_url)
            if response.status_code == 200:
                with open(ruta_main+'/'+nombre_cancion+'.jpg', "wb") as f:
                    f.write(response.content)
                print(f"Miniatura guardada en: {ruta_main}")
            else:
                print("Error al descargar la miniatura.")
            #FINAL CACHO QUE DESCARGA IMAGENES

            yt_audio = yt.streams.filter(only_audio=True).first()
            if yt_audio is None:
                print("ERROR: No se encontró stream de audio para el video para descarga.")
                mensaje_error(ventana_secundaria, "No se encontró audio para este video.")
                return
            
            archivo_audio_path = yt_audio.download(output_path=ruta_main, filename=nombre_cancion+".mp3")
            
        elif En_biblioteca is True:
            validador=True
            nombreExistente = hoja["B"+str(linea_clave)].value
            nombreExistente=str(nombreExistente)
            archivo_audio_path = os.path.join(ruta_main, nombreExistente) # Full path to existing MP3
            
            if not os.path.exists(archivo_audio_path):
                print(f"ADVERTENCIA: La canción '{nombreExistente}' no se encontró en el disco. Intente descargarla de nuevo.")
                mensaje_error(ventana_secundaria, "La canción no se encontró en su biblioteca local.")
                validador = False
        
        if validador and os.path.exists(archivo_audio_path):
            administrar_audio(archivo_audio_path,etiqueta_cancion_actual)
        elif validador:
            print("No se pudo reproducir la canción descargada/existente.")
        mostrar_canciones(importar_canciones(lista),frame_canciones_placeholder,etiqueta_cancion_actual,frame_playlists_placeholder,frame_playlist_canciones_placeholder) # Use placeholder


def mostrar_playlists(frame_playlists_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder): #Tiene Color
    global playlists_registro
    importar_playlists()
    print(f"mostrar_playlists llamado con: {playlists_registro}") # <-- IMPORTE
    limpiar_frame(frame_playlists_placeholder)

    # Ensure the inner frame's column 0 expands with content
    frame_playlists_placeholder.grid_columnconfigure(0, weight=1) # Play button column, fixed size
    frame_playlists_placeholder.grid_columnconfigure(1, weight=1) # Song label column, expands
    frame_playlists_placeholder.grid_columnconfigure(2, weight=0) # Song label column, expands
    i=0
    for playlist in playlists_registro:
        if isinstance(playlist, list) and len(playlist) > 0:
            playlist_nombre = playlist[0]
        else:
            print(f"Elemento de playlist inesperado o vacío: {playlist}")
        playlist_nombre = playlist[0]

        song_row_frame = tk.Frame(frame_playlists_placeholder, bg="#008B7E") # Greenish background for song rows
        song_row_frame.grid(row=i, column=0, columnspan=2, sticky="ew", padx=5, pady=3) # Span all 2
        
        #Solo van a ser 3
        song_row_frame.grid_columnconfigure(0, weight=0) # Para el boton reproducir
        song_row_frame.grid_columnconfigure(1, weight=1) # Para el nombre de la playlist
        song_row_frame.grid_columnconfigure(2, weight=0) # Para el boton eliminar

        boton_reproducir_playlist = tk.Button(song_row_frame, text="▶", foreground="#FFFFFF", bg="#008B7E",
                                    font="Corbel 10", justify="center", activebackground="#0273A7",
                                    activeforeground="#FFFFFF", 
                                    command=lambda a=playlist: reproducir_playlist_random(a,etiqueta_cancion_actual,frame_playlist_canciones_placeholder),
                                    width=2) # Fixed width for consistent button size
        
        etiqueta_playlist = tk.Label(song_row_frame, text=playlist_nombre, foreground="#FFFFFF", bg="#008B7E",
                                    justify="left", font="Corbel 16", anchor="w", wraplength=400) # Added wraplength

        boton_eliminar_playlist = tk.Button(song_row_frame, text="X", foreground="#FFFFFF", bg="#008B7E",
                                    font="Corbel 10", justify="center", activebackground="#D80404",
                                    activeforeground="#FFFFFF", width=2,
                                    command=lambda a=playlist : eliminar_playlist(a,frame_playlists_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder)
                                    ) 

        boton_reproducir_playlist.grid(row=0, column=0, padx=(5,2), pady=2, sticky="w")
        etiqueta_playlist.grid(row=0, column=1, padx=2, pady=2, sticky="ew")
        boton_eliminar_playlist.grid(row=0, column=2, padx=2, pady=2, sticky="ew")
        i=i+1

    frame_playlists_placeholder.update_idletasks() # Ensure widgets are drawn and sizes calculated
    frame_playlists_placeholder.master.config(scrollregion=frame_playlists_placeholder.master.bbox("all"))

    
def mostrar_canciones(canciones_registro, frame_canciones_placeholder, etiqueta_cancion_actual, frame_playlists_placeholder,frame_playlist_canciones_placeholder):#Tiene Color
    limpiar_frame(frame_canciones_placeholder) # Clean the inner frame
    # Ensure the inner frame's column 0 expands with content
    frame_canciones_placeholder.grid_columnconfigure(0, weight=1) # Play button column, fixed size
    frame_canciones_placeholder.grid_columnconfigure(1, weight=1) # Song label column, expands
    frame_canciones_placeholder.grid_columnconfigure(2, weight=0) # Add to queue button column, fixed size
    frame_canciones_placeholder.grid_columnconfigure(3, weight=0) # Add to delete button column

    i=0
    print(canciones_registro)
    for cancion_nombre_mp3 in canciones_registro.values(): # Iterate over values (song names like "Kakegurui_Twin.mp3")
        # Ensure the full path to the audio file
        archivo_audio = os.path.join(ruta_main, cancion_nombre_mp3)
        display_name = cancion_nombre_mp3.replace('.mp3', '')

        # Create a sub-frame for each song row to ensure proper alignment and background
        song_row_frame = tk.Frame(frame_canciones_placeholder, bg="#008B7E") # Greenish background for song rows
        # Place this sub-frame in the grid of frame_canciones_placeholder
        song_row_frame.grid(row=i, column=0, columnspan=4, sticky="ew", padx=5, pady=3) # Span all 3 columns #Modifique esto

        # Configure columns within the song_row_frame for its widgets
        song_row_frame.grid_columnconfigure(0, weight=0) # Para el boton reproducir
        song_row_frame.grid_columnconfigure(1, weight=1) # Para el nombre de la canción
        song_row_frame.grid_columnconfigure(2, weight=0) # Para agregar a la cola
        song_row_frame.grid_columnconfigure(3, weight=0) # Para agregar 
        song_row_frame.grid_columnconfigure(4, weight=0) # Para borrar canción


        boton_reproducir_canciones = tk.Button(song_row_frame, text="▶", foreground="#FFFFFF", bg="#008B7E",
                                               font="Corbel 14", justify="center", activebackground="#0273A7",
                                               activeforeground="#FFFFFF", 
                                               command=lambda a=archivo_audio: administrar_audio(a,etiqueta_cancion_actual),
                                               width=2) # Fixed width for consistent button size
        
        etiqueta_cancion = tk.Label(song_row_frame, text=display_name, foreground="#FFFFFF", bg="#008B7E",
                                    justify="left", font="Corbel 16", anchor="w", wraplength=400) # Added wraplength
        
        boton_agregar_cola = tk.Button(song_row_frame, text="➥", foreground="#FFFFFF", bg="#008B7E",
                                        font="Corbel 10", justify="center", activebackground="#02A796",
                                        activeforeground="#FFFFFF", width=2,
                                        command=lambda a=archivo_audio: agregar_cola(a)) # Fixed width for consistent button size

        boton_eliminar_cancion = tk.Button(song_row_frame, text="x", foreground="#FFFFFF", bg="#008B7E",
                                        font="Corbel 10", justify="center", activebackground="#D80404",
                                        activeforeground="#FF0000", width=2,
                                        command=lambda a=archivo_audio: eliminar_cancion(canciones_registro,a,frame_canciones_placeholder,etiqueta_cancion_actual,frame_playlists_placeholder,frame_playlist_canciones_placeholder)) # Fixed width for consistent button size
        
        boton_agregar_playlist = tk.Button(song_row_frame, text="+", foreground="#FFFFFF", bg="#008B7E",
                                        font="Corbel 10", justify="center", activebackground="#02A796",
                                        activeforeground="#FFFFFF", width=2,
                                        command=lambda a=cancion_nombre_mp3: menu_canciones_playlist(a, etiqueta_cancion_actual, frame_playlists_placeholder,frame_playlist_canciones_placeholder)
                                        ) # Fixed width for consistent button size
        
        # Place widgets within the song_row_frame's grid
        boton_reproducir_canciones.grid(row=0, column=0, padx=(5,2), pady=2, sticky="w")
        etiqueta_cancion.grid(row=0, column=1, padx=2, pady=2, sticky="ew")
        boton_agregar_cola.grid(row=0, column=2, padx=(2,5), pady=2, sticky="e")
        boton_agregar_playlist.grid(row=0, column=3, padx=(2,5), pady=2, sticky="e")
        boton_eliminar_cancion.grid(row=0, column=4, padx=(2,5), pady=2, sticky="e")
        
        i=i+1

    # After updating the content, ensure the scroll region is updated
    # This bind is essential for the scrollbar to function correctly
    frame_canciones_placeholder.update_idletasks() # Ensure widgets are drawn and sizes calculated
    frame_canciones_placeholder.master.config(scrollregion=frame_canciones_placeholder.master.bbox("all"))

def reproducir_playlist_random(playlist, etiqueta_cancion_actual,frame_playlist_canciones_placeholder):
    global cola_reproduccion
    if len(playlist) < 2:
        print("Playlist vacía")
        mensaje_error(ventana_principal, "Playlist vacía")
        return
    
    mostrar_playlist_canciones(playlist,frame_playlist_canciones_placeholder,etiqueta_cancion_actual)

    canciones = playlist[1:]  # Ignora el nombre de la playlist
    randomizador = r.randint(0, 1)

    if randomizador == 1:
        r.shuffle(canciones)      # Reordenar aleatoriamente
        cola_reproduccion = Queue()  # Vacía la cola anterior
        for cancion in canciones:
            archivo_audio = os.path.join(ruta_main, cancion)
            cola_reproduccion.enqueue(archivo_audio)
        print("Playlist RANDOM a la cola:", canciones)
    else:
        insertionSort(canciones)
        print('PRESTA ATENCION',canciones)
        for cancion in canciones:
            archivo_audio = os.path.join(ruta_main, cancion)
            cola_reproduccion.enqueue(archivo_audio)
        print("Playlist ALFABETICA a la cola:", canciones)
    # Reproducir la primera de la cola
    if cola_reproduccion.size() > 0:
        administrar_audio(cola_reproduccion.dequeue(), etiqueta_cancion_actual)


##AQUI ESTA EL ALGORITMO de insertionSort
def insertionSort(lista):
    for index in range(1, len(lista)):
        actual = lista[index]  # Elemento actual para colocarlo en la posición correcta
        posicion = index  # Posición de dicho elemento
        print("Valor a ordenar = {}".format(actual))
        while posicion > 0 and lista[posicion - 1] > actual:
            lista[posicion] = lista[posicion - 1]  # Mueve el elemento a la derecha
            posicion = posicion - 1  # Mueve la posición a la izquierda
        lista[posicion] = actual  # Coloca el valor actual en la posición correcta
        print(lista)
        print()
    return lista


def mostrar_playlist_canciones(playlist,frame_playlist_canciones_placeholder,etiqueta_cancion_actual):#Tiene Color
    limpiar_frame(frame_playlist_canciones_placeholder)
    print("Hola si entro")
    canciones = playlist[1:]
    print(canciones)
    frame_playlist_canciones_placeholder.grid_columnconfigure(0, weight=0) 
    frame_playlist_canciones_placeholder.grid_columnconfigure(1, weight=1) 
    frame_playlist_canciones_placeholder.grid_columnconfigure(2, weight=0) 
    i = 0
    for cancion in canciones:
        print(i,cancion)
        display_name = cancion.replace(".mp3", "")

        song_row_frame = tk.Frame(frame_playlist_canciones_placeholder, bg="#008B7E") # Greenish background for song rows
        song_row_frame.grid(row=i, column=0, columnspan=2, sticky="ew", padx=5, pady=3) # Span all 2

        song_row_frame.grid_columnconfigure(0, weight=0) # Para el boton reproducir
        song_row_frame.grid_columnconfigure(1, weight=1) # Para el nombre de la canción
        song_row_frame.grid_columnconfigure(2, weight=0) # Para eliminar cancion de playlist

        boton_reproducir_cancion_playlist = tk.Button(song_row_frame, text="▶", foreground="#FFFFFF", bg="#008B7E",
                                    font="Corbel 10", justify="center", activebackground="#0273A7",
                                    activeforeground="#FFFFFF", 
                                    command=lambda a=os.path.join(ruta_main, cancion): administrar_audio(a,etiqueta_cancion_actual),
                                    width=2) 
        
        etiqueta_cancion = tk.Label(song_row_frame, text=display_name, foreground="#FFFFFF", bg="#008B7E",
                                    justify="left", font="Corbel 10", anchor="w", wraplength=400) # Added wraplength

        boton_eliminar_cancion_playlist = tk.Button(song_row_frame, text="X", foreground="#FFFFFF", bg="#008B7E",
                                    font="Corbel 10", justify="center", activebackground="#D80404",
                                    activeforeground="#FFFFFF", width=2,
                                    command=lambda a=playlist,b=cancion : (eliminar_cancion_playlist(a,b),mostrar_playlist_canciones(a,frame_playlist_canciones_placeholder,etiqueta_cancion_actual)),
                                    ) 
        
        boton_reproducir_cancion_playlist.grid(row=0, column=0, padx=(5,2), pady=2, sticky="w")
        etiqueta_cancion.grid(row=0, column=1, padx=2, pady=2, sticky="ew")
        boton_eliminar_cancion_playlist.grid(row=0, column=2, padx=2, pady=2, sticky="ew")
        i=i+1
    frame_playlist_canciones_placeholder.update_idletasks() # Ensure widgets are drawn and sizes calculated
    frame_playlist_canciones_placeholder.master.config(scrollregion=frame_playlist_canciones_placeholder.master.bbox("all"))

#Si hay un error es aqui
def eliminar_cancion_playlist(playlist,cancion):
    global hoja
    hoja = archivo[usuario_activo]
    j=1
    print(playlist)
    print(cancion)
    while(hoja["C"+str(j)].value is not None):
        if hoja["C" + str(j)].value == str(playlist):
            lista_aux = ast.literal_eval(hoja["C"+str(j)].value)
            if cancion in lista_aux:
                print("Aqui está la cansion")
                playlist.remove(cancion)
                hoja["C" + str(j)].value = str(playlist)
        j = j+1
    archivo.save("canciones.xlsx") #GUARDA CAMBIOS HECHOS EN EXCEL
    importar_playlists()

# --- Inicio de la modificación en ventana_reproductor ---
def ventana_reproductor():#Tiene Color
    global ventana_principal
    global progress_scale
    global update_progress_job
    global imagen_label
    global playlists_registro

    for widget in ventana_principal.winfo_children():
        widget.destroy()

    ventana_principal.title("MUZIKAI - Reproductor") # Updated title
    ventana_principal.geometry("1200x800")
    ventana_principal.config(bg = "#029485")
    ventana_principal.deiconify()
    ventana_principal.resizable(False,False)
    canciones_registro = {}
    playlists_registro = []
    importar_canciones(canciones_registro)
    importar_playlists()

    etiqueta1 = tk.Label (ventana_principal, #Mensaje de ese
                            text="Ingrese un link de youtube:",
                            foreground="#FFFFFF",
                            bg="#029485",
                            font="Corbel 18",
                            )

    entrada_barra = tk.Entry (ventana_principal, #Entrada para el link de yt
                            foreground="#FFFFFF",
                            bg="#038376",
                            width="30",
                            font="Corbel 20",
                            )

    boton_busqueda = tk.Button (ventana_principal, #Boton para el link de yt
                            text="Buscar",
                            foreground="#FFFFFF",
                            bg="#008B7E",
                            height="1",
                            font="Corbel 12",
                            justify="center",
                            activebackground="#14978A",
                            activeforeground="#FFFFFF",
                            command=lambda: menu_descarga(entrada_barra,canciones_registro,ventana_principal,frame_canciones_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder,frame_playlists_placeholder) # Pass frame_canciones_placeholder
                            )


    frame_playlists = tk.Frame(ventana_principal, #Frame para las playlists
                            bg="#006B60",
                            width=250,
                            height=520
                            )

    frame_controles = tk.Frame(ventana_principal, #Frame para los coltroles (pausa, sig, barra)
                            bg="#057062",
                            height=110
                            )

    frame_cancion_lista = tk.Frame(ventana_principal, #Frame para imagen y nombre
                            bg="#006B60",
                            width=250,
                            height=300
                            )


    frame_canciones_contenedor = tk.Frame(ventana_principal, #Frame que tierne el canva con canciones y el scrollbar
                            bg="#006B60", 
                            width=530,
                            height=405
                            )
    
    frame_playlist_canciones = tk.Frame(ventana_principal, #Frame que tierne el canva con canciones y el scrollbar
                            bg="#006B60", 
                            width=250,
                            height=230
                            )
    



# --- Scrolling Implementation for frame_canciones ---
    canvas = tk.Canvas(frame_canciones_contenedor,  #Canva para contener canciones dentro de frame_canciones_contenedor
                            bg="#006B60", 
                            highlightthickness=0
                            ) 
    
    scrollbar = ttk.Scrollbar(frame_canciones_contenedor,  #Scrollbar dentro de frame_canciones_contenedor vinculado al canvas
                            orient="vertical", 
                            command=canvas.yview
                            )
    # Create an inner frame within the Canvas to hold the song entries, and this is where your grid layout for songs will go.
    frame_canciones_placeholder = tk.Frame(canvas, 
                            bg="#006B60"
                            ) 

    # Add the inner frame to a window in the canvas
    canvas_window_id = canvas.create_window((0, 0), 
                            window=frame_canciones_placeholder, 
                            anchor="nw")

    # Function to adjust the width of the frame_canciones_placeholder window within the canvas and to update the scrollregion.
    def on_canvas_configure(event):
        canvas.itemconfig(canvas_window_id, width=event.width) # Update the width of the window item to match the canvas width
        canvas.configure(scrollregion=canvas.bbox("all"))# Update the scrollregion. This is critical for scrolling to work.

    
    canvas.bind('<Configure>', on_canvas_configure) # Bind the canvas's <Configure> event to this function
    # Bind the inner frame's <Configure> event to update the scrollregion. This is important when content within `frame_canciones_placeholder` changes size.
    frame_canciones_placeholder.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
# --- End of Scrolling Implementation ---


# --- Copia del scrolling pero para frame_playlist ---
    canvas2 = tk.Canvas(frame_playlists,  #Canva para contener canciones dentro de frame_canciones_contenedor
                            bg="#006B60", 
                            width=235,
                            highlightthickness=0
                            ) 
    
    scrollbar2 = ttk.Scrollbar(frame_playlists,  #Scrollbar dentro de frame_canciones_contenedor vinculado al canvas
                            orient="vertical", 
                            command=canvas2.yview
                            )
    # Create an inner frame within the Canvas to hold the song entries, and this is where your grid layout for songs will go.
    frame_playlists_placeholder = tk.Frame(canvas2, 
                            width=200,
                            bg="#006B60"
                            ) 

    # Add the inner frame to a window in the canvas
    canvas2_window_id = canvas2.create_window((0, 0), 
                            window=frame_playlists_placeholder,
                            width=200, 
                            anchor="nw")

    # Function to adjust the width of the frame_canciones_placeholder window within the canvas and to update the scrollregion.
    def on_canvas_configure2(event):
        canvas2.itemconfig(canvas2_window_id, width=event.width) # Update the width of the window item to match the canvas width
        canvas2.configure(scrollregion=canvas2.bbox("all"))# Update the scrollregion. This is critical for scrolling to work.

    
    canvas2.bind('<Configure>', on_canvas_configure2) # Bind the canvas's <Configure> event to this function
    # Bind the inner frame's <Configure> event to update the scrollregion. This is important when content within `frame_canciones_placeholder` changes size.
    frame_playlists_placeholder.bind("<Configure>", lambda o: canvas2.configure(scrollregion=canvas2.bbox("all")))
# --- End of Scrolling Implementation ---


# --- Copia del scrolling pero para frame_playlist_canciones ---
    canvas3 = tk.Canvas(frame_playlist_canciones,  #Canva para contener canciones dentro de frame_canciones_contenedor
                            bg="#006B60", 
                            width=235,
                            highlightthickness=0
                            ) 
    
    scrollbar3 = ttk.Scrollbar(frame_playlist_canciones,  #Scrollbar dentro de frame_canciones_contenedor vinculado al canvas
                            orient="vertical", 
                            command=canvas3.yview
                            )
    # Create an inner frame within the Canvas to hold the song entries, and this is where your grid layout for songs will go.
    frame_playlist_canciones_placeholder = tk.Frame(canvas3, 
                            width=200,
                            bg="#006B60"
                            ) 

    # Add the inner frame to a window in the canvas
    canvas3_window_id = canvas3.create_window((0, 0), 
                            window=frame_playlist_canciones_placeholder,
                            width=200, 
                            anchor="nw")

    # Function to adjust the width of the frame_canciones_placeholder window within the canvas and to update the scrollregion.
    def on_canvas_configure3(event):
        canvas3.itemconfig(canvas3_window_id, width=event.width) # Update the width of the window item to match the canvas width
        canvas3.configure(scrollregion=canvas3.bbox("all"))# Update the scrollregion. This is critical for scrolling to work.

    
    canvas3.bind('<Configure>', on_canvas_configure3) # Bind the canvas's <Configure> event to this function
    # Bind the inner frame's <Configure> event to update the scrollregion. This is important when content within `frame_canciones_placeholder` changes size.
    frame_playlist_canciones_placeholder.bind("<Configure>", lambda o: canvas3.configure(scrollregion=canvas3.bbox("all")))
# --- End of Scrolling Implementation ---



    etiqueta_playlist_actual = tk.Label (frame_playlist_canciones,
                            text="Canciones playlist actual",
                            foreground="#FFFFFF",
                            bg="#006B60",
                            font="Corbel 16",
                            justify="left",
                            wraplength=250
                          )


    etiqueta_cancion_actual = tk.Label (frame_cancion_lista,
                            text="",
                            foreground="#FFFFFF",
                            bg="#006B60",
                            font="Corbel 18",
                            justify="left",
                            wraplength=250
                          )
    
    boton_reproducir = tk.Button (frame_controles,
                            text="▶",
                            foreground="#FFFFFF",
                            bg="#008B7E",
                            font="Corbel 16",
                            activebackground="#02A796",
                            activeforeground="#FFFFFF",
                            command=lambda: alt_pausa()
                            )
    
    boton_siguiente = tk.Button (frame_controles,
                            text="»",
                            foreground="#FFFFFF",
                            bg="#008B7E",
                            font="Corbel 16",
                            justify="center",
                            activebackground="#02A796",
                            activeforeground="#FFFFFF",
                            command=lambda: reproducir_siguiente(etiqueta_cancion_actual) ############
                            )    

    boton_agregar_playlist = tk.Button (ventana_principal,
                            text="Agregar playlist",
                            foreground="#FFFFFF",
                            bg="#008B7E",
                            font="Corbel 12",
                            justify="center",
                            activebackground="#02A796",
                            activeforeground="#FFFFFF",
                            command=lambda: menu_agregar_playlist(ventana_principal,frame_playlists_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder)
                            )    
    boton_admin_panel = tk.Button (ventana_principal,
                            text="⛭Admin",
                            foreground="#FFFFFF",
                            bg="#008B7E",
                            font="Corbel 12",
                            justify="center",
                            activebackground="#02A796",
                            activeforeground="#FFFFFF",
                            command=lambda: menu_usuarios()
                            )   
    
    progress_scale = tk.Scale(frame_controles,
                            from_=0, # Valor mínimo (0 segundos)
                            to=3,  # Se actualizará con la duración total
                            orient=tk.HORIZONTAL,
                            length=500, # Ancho de la barra
                            showvalue=0, # No mostrar el valor numérico directamente en la barra
                            bg="#10A792", # Mismo color de fondo que frame_controles
                            fg="white",
                            highlightthickness=0, # Eliminar el borde del scale
                            troughcolor="#0F9E90", # Color de la "pista" del control deslizante
                            activebackground="#02A796",
                            command=on_scale_move # Función que se llama cuando el usuario mueve la barra
                    )
    
#Para la cola
    verificar_cola(etiqueta_cancion_actual)


#Configs
    canvas.configure(yscrollcommand=scrollbar.set)
    canvas2.configure(yscrollcommand=scrollbar2.set)
    canvas3.configure(yscrollcommand=scrollbar3.set)

    
#Evitar que se propaguen
    frame_canciones_contenedor.pack_propagate(False) 
    frame_cancion_lista.pack_propagate(False)
    frame_playlists.pack_propagate(False)
    frame_controles.pack_propagate(False)
    frame_playlist_canciones.pack_propagate(False)

#Llamar a todo
    print(admin)
    if admin == "admin":
        boton_admin_panel.place(x=1070,y=15)

    etiqueta1.pack(pady=20)
    entrada_barra.pack()
    boton_busqueda.pack(pady=6)
    etiqueta_playlist_actual.pack(pady=3)
    
    canvas.pack(side="left", fill="both", expand=True) ##########
    scrollbar.pack(side="right", fill="y")
    canvas2.pack(side="left", fill="both", expand=True) ##########
    scrollbar2.pack(side="right", fill="y")
    canvas3.pack(side="left", fill="both", expand=True) ##########
    scrollbar3.pack(side="right", fill="y")
    frame_canciones_contenedor.place(x=330, y=225)
    frame_playlists.place(x=40,y=110)
    frame_cancion_lista.place(x=900,y=60)
    frame_playlist_canciones.place(x=900,y=400)

    #Añadido para imagenes
    # Esto va justo después de frame_cancion_lista.place(...)
    imagen_label = tk.Label(frame_cancion_lista, bg="#006B60")
    imagen_label.pack(pady=10)
    #Fin de añadido para imagenes

    etiqueta_cancion_actual.pack()
    frame_controles.pack(side=tk.BOTTOM,fill=tk.X)
    boton_reproducir.pack()
    boton_siguiente.place(x=650)
    boton_agregar_playlist.place(x=70,y=60)
    progress_scale.pack(pady=5, fill=tk.X, padx=10) # Empaquetar la barra de progreso

    mostrar_canciones(canciones_registro, frame_canciones_placeholder, etiqueta_cancion_actual, frame_playlists_placeholder,frame_playlist_canciones_placeholder)
    mostrar_playlists(frame_playlists_placeholder,etiqueta_cancion_actual,frame_playlist_canciones_placeholder)


# --- Fin de la modificación en ventana_reproductor ---

def actualizar_imagen(ruta_imagen):
    global imagen_label
    try:
        imagen = Image.open(ruta_imagen)
        imagen = imagen.resize((240, 240))
        # Definimos el área visible (por ejemplo, 250x250 desde la esquina superior izquierda)
        area_visible = (0,30,240,210)  # (left, upper, right, lower)
        imagen_crop = imagen.crop(area_visible)
        # Convertimos a imagen de Tkinter y la colocamos
        imagen_tk = ImageTk.PhotoImage(imagen_crop)
        imagen_label.config(image=imagen_tk)
        imagen_label.image = imagen_tk
    except FileNotFoundError:
        print("No hay imagen mano jiji")
        imagen_label.config(image='') # Quita la imagen

def crear_cuenta():#Tiene Color
    global ventana_descarga
    if ventana_descarga is not None and ventana_descarga.winfo_exists():
        ventana_descarga.focus()
        return
    ventana_secundaria = tk.Toplevel(ventana_principal)
    ventana_descarga = ventana_secundaria
    ventana_secundaria.title("MUZIKAI - Crear Cuenta") # Updated title
    ventana_secundaria.geometry("800x500")
    ventana_secundaria.resizable(False,False)
    ventana_secundaria.config(bg = "#02695F")
    ventana_descarga.protocol("WM_DELETE_WINDOW", lambda: cerrar_ventana_descarga())

    etiqueta_c = tk.Label (ventana_secundaria,
                           text="Creación de cuenta",
                           foreground="#FFFFFF",
                           bg="#02695F",
                           width="19",
                           height="1",
                           font="Corbel 50",
                           relief="groove"
                           )
    
    etiqueta1 = tk.Label (ventana_secundaria,
                           text="Nombre de usuario:",
                           foreground="#FFFFFF",
                           bg="#02695F",
                           font="Corbel 18",
                           )

    entrada1 = tk.Entry (ventana_secundaria,
                           foreground="#FFFFFF",
                           bg="#00574E",
                           width="30",
                           font="Corbel 20",
                           )

    etiqueta2 = tk.Label (ventana_secundaria,
                           text="Contraseña:",
                           foreground="#FFFFFF",
                           bg="#02695F",
                           font="Corbel 18",
                           )

    entrada2 = tk.Entry (ventana_secundaria,
                           foreground="#FFFFFF",
                           bg="#00574E",
                           width="30",
                           font="Corbel 20",
                           show="*"
                           )

    boton_ver = tk.Button(ventana_secundaria,
                           text="Ver",
                           command=lambda: alternar_visibilidad_contrasena(entrada2),
                           bg="#02695F",
                           foreground="#FFFFFF",
                           font="Corbel 10",
                           )


    boton1 = tk.Button (ventana_secundaria,
                           text="Crear cuenta",
                           foreground="#FFFFFF",
                           bg="#008B7E",
                           width="25",
                           height="1",
                           font="Corbel 30",
                           justify="center",
                           activebackground="#02A796",
                           activeforeground="#FFFFFF",
                           command=lambda: agregar_sesion(entrada1,entrada2,ventana_secundaria,boton1)
                           )

    etiqueta_c.pack(pady=20)
    etiqueta1.pack()
    entrada1.pack()
    etiqueta2.pack()
    entrada2.pack()
    boton_ver.pack()
    boton1.pack(pady=20)


def inicio_sesion():#Tiene Color
    global usuario_activo
    global ruta_main
    global ventana_principal # Asegurar que accedemos a la global
    global hoja
    global admin
    usuario = entrada1.get()
    token = entrada2.get()
    hoja=archivo["USUARIOS"]

    encontrado = False
    for i in range(1,hoja.max_row+1):
        if usuario == hoja["A"+str(i)].value and token== hoja["B"+str(i)].value:
            encontrado=True
            admin = hoja["C"+str(i)].value

    if encontrado:
        etiqueta = tk.Label (ventana_principal,
                             text="Iniciando sesión",
                             foreground="#014942",
                             bg="#014942",
                             font="Corbel 18",
                             )
        etiqueta.pack()
        boton1.config(state=tk.DISABLED)
        usuario_activo = usuario
        ruta_main=os.path.abspath(usuario_activo.upper())

        ventana_principal.withdraw()
        ventana_principal.after(3000, ventana_reproductor)

        print("Id es",usuario,"y token es",token)
    else:
        print("Datos erroneos")
        mensaje_error(ventana_principal,"Datos no validos")
    return usuario


def agregar_sesion(entradan1,entradan2,ventana,boton):#Tiene Color
    global hoja
    usuario = entradan1.get()
    token = entradan2.get()
    hoja=archivo["USUARIOS"]
    repetido=False
    if hoja["A1"].value != None:
        for i in range(1,hoja.max_row+1):
            if (usuario.strip()).upper()==(hoja["A"+str(i)].value).upper():
                repetido=True
                break

    if (usuario == "" or token == ""):
        mensaje_error(ventana,"Datos no validos")

    elif repetido:
        mensaje_error(ventana,"Usuario ya tomado, use otro")

    else:
        etiqueta = tk.Label (ventana,
                    text="Cuenta creada, inicie sesión",
                    foreground="#014942",
                    bg="#014942",
                    font="Corbel 18",
                    )

        usuario_completo=[usuario,token,'usuario']
        if hoja["A1"].value is None:
            hoja["A1"].value=usuario
            hoja["B1"].value=token
            hoja["C1"].value='usuario'
        else:
            hoja.append(usuario_completo)
        archivo.create_sheet(title=usuario)
        archivo.save("canciones.xlsx")
        os.makedirs(usuario.upper(), exist_ok=True) # Use exist_ok=True to prevent error if folder exists


        etiqueta.pack()
        boton.config(state=tk.DISABLED)
        ventana.after(3000, lambda: cerrar_ventana(ventana))
        print("Id es",usuario,"y token es",token)

def eliminar_sesion(nombre_usuario):
    global archivo, ruta_main
    global current_vlc_player
    hoja = archivo["USUARIOS"]
    nombre_usuario_upper = nombre_usuario.strip().upper()
    encontrado = False

    # Buscar el usuario en la hoja "USUARIOS" y eliminar su fila
    for i in range(1, hoja.max_row + 1):
        celda = hoja[f"A{i}"]
        if celda.value and celda.value.strip().upper() == nombre_usuario_upper:
            hoja.delete_rows(i, 1)
            encontrado = True
            break

    if not encontrado:
        print("Usuario no encontrado.")
        return

    # Eliminar la hoja con el nombre del usuario
    if nombre_usuario in archivo.sheetnames:
        std = archivo[nombre_usuario]
        archivo.remove(std)
    else:
        print("Hoja del usuario no encontrada.")

    # Guardar cambios en el Excel
    archivo.save("canciones.xlsx")
    if current_vlc_player:
        current_vlc_player.stop()
        current_vlc_player.release()
        current_vlc_player = None
    if imagen_label:
        imagen_label.config(image='')
        imagen_label.image = None
    # Eliminar carpeta del usuario
    carpeta_usuario = os.path.abspath(nombre_usuario_upper)
    print (carpeta_usuario)
    if os.path.exists(carpeta_usuario):
        try:
            shutil.rmtree(carpeta_usuario)
            print("Carpeta del usuario eliminada.")
        except Exception as e:
            print(f"No se pudo eliminar la carpeta: {e}")
    else:
        print("Carpeta del usuario no encontrada.")


def mensaje_error(ventana,mensaje):#Tiene Color
        etiqueta = tk.Label (ventana,
                             text=mensaje,
                             foreground="#FF0000",
                             bg="#025F56",
                             font="Corbel 18",
                             )
        etiqueta.pack(pady=2)
        ventana.after(1500, lambda: destruir_etiqueta_si_existe(etiqueta))


def formato_tiempo(segundos_totales): #Para dale formato al tiempo

    if not isinstance(segundos_totales, int) or segundos_totales < 0:
        raise ValueError("El segundos_totales debe ser un entero no negativo.")

    horas = segundos_totales // 3600
    minutos = (segundos_totales % 3600) // 60
    segundos = segundos_totales % 60

    minutos_formato = f"{minutos:02d}"
    segundos_formato = f"{segundos:02d}"

    if horas > 0:
        horas_formato = f"{horas:02d}"
        return f"{horas_formato}:{minutos_formato}:{segundos_formato}"
    else:
        return f"{minutos_formato}:{segundos_formato}"


#---------Para la barra de carga (mayormente IA)------------------------------------------------------------------------------------

# Nueva función para manejar el movimiento de la barra de progreso por el usuario
def cancel_update_progress():
    """Cancela la tarea programada para actualizar la barra de progreso."""
    global update_progress_job
    global ventana_principal # Necesaria para after_cancel
    if update_progress_job:
        ventana_principal.after_cancel(update_progress_job)
        update_progress_job = None

def update_progress_bar():
    """Actualiza la posición de la barra de progreso en segundos y se reprograma."""
    global current_vlc_player
    global progress_scale
    global update_progress_job
    global ventana_principal # Necesaria para after
    global program_is_setting_scale

    if current_vlc_player:
        if current_vlc_player.is_playing():
            current_time_ms = current_vlc_player.get_time()
            total_length_ms = current_vlc_player.get_length()

            if total_length_ms > 0:
                current_time_seconds = current_time_ms // 1000
                if progress_scale:
                    program_is_setting_scale = True
                    progress_scale.set(current_time_seconds)
                    program_is_setting_scale = False
            elif progress_scale: # Si está reproduciendo pero la duración es 0 o inválida
                program_is_setting_scale = True
                progress_scale.set(0)
                program_is_setting_scale = False
        
        # Reprogramar la próxima actualización si el reproductor aún existe
        # Considera aumentar el intervalo si 100ms sigue siendo problemático (ej. 200 o 250)
        update_progress_job = ventana_principal.after(100, update_progress_bar)
    else:
        cancel_update_progress()


def setup_progress_bar(current_vlc_player):
    """Configura la barra de progreso para el reproductor actual (en segundos) e inicia las actualizaciones."""
    global progress_scale
    global update_progress_job # Necesario para cancel_update_progress y update_progress_bar
    global program_is_setting_scale

    if current_vlc_player:
        total_length_ms = current_vlc_player.get_length()
        total_duration_seconds = total_length_ms // 1000
        #print("---------",total_duration_seconds)

        if progress_scale: # Asegurarse de que progress_scale ha sido creado
            if total_duration_seconds > 0:
                print("\n",total_duration_seconds,"\n")
                progress_scale.config(to=total_duration_seconds)
            else:
                # Si la duración es 0 o inválida, usar un valor predeterminado para 'to' y advertir. La barra mostrará 0.
                progress_scale.config(to=100) # Un valor predeterminado, podría ser cualquier cosa
                print("ADVERTENCIA: Duración del audio es 0 o inválida. La barra de progreso podría no reflejar la duración real.")
            
            program_is_setting_scale = True
            progress_scale.set(0) # Inicializar en 0 segundos
            program_is_setting_scale = False
        
        # Siempre iniciar/reiniciar el bucle de actualización si hay un reproductor.
        cancel_update_progress()
        update_progress_bar()
    else:
        if progress_scale:
            program_is_setting_scale = True
            progress_scale.set(0)
            program_is_setting_scale = False
        cancel_update_progress()

def on_scale_move(value_str_seconds):
    """Maneja el evento cuando el usuario mueve la barra de progreso (valor en segundos)."""
    global current_vlc_player
    global ventana_principal # Necesaria para after
    global program_is_setting_scale

    if program_is_setting_scale:
        # No hacer nada si el cambio de valor fue activado por progress_scale.set() desde update_progress_bar()
        return

    if current_vlc_player and \
        (current_vlc_player.is_playing() or current_vlc_player.get_state() == vlc.State.Paused):
        
        total_length_ms = current_vlc_player.get_length()
        if total_length_ms > 0:

            new_time_ms = int(value_str_seconds)*1000 # Convertir segundos a milisegundos
            
            #Si por cualquier cosa se pasa de la duracion total
            if new_time_ms > total_length_ms:
                new_time_ms = total_length_ms
            if new_time_ms < 0:
                new_time_ms = 0
            
            was_playing_before_seek = current_vlc_player.is_playing()
            
            cancel_update_progress() 

#Para poner los segundos de las canciones en la pantalla ------------------------------------
            #print(value_str_seconds)
            segundo_cancion = tk.Label(ventana_principal,
                            text=formato_tiempo(int(value_str_seconds)),
                            foreground="#FFFFFF",
                            bg="#02695F",
                            font="Corbel 18",
                            )

            segundo_total = tk.Label(ventana_principal,
                            text=formato_tiempo(int(total_length_ms/1000)),
                            foreground="#FFFFFF",
                            bg="#02695F",
                            font="Corbel 18",
                            )
            
            segundo_cancion.place(x=30,y=700)
            segundo_total.place(x=1100,y=700)
#------------------------------------------------------------------------------------------
            
            if abs(current_vlc_player.get_time() - new_time_ms)>1500: #ESTA LINEA LA HICE YO PROFE E, LE TUVE Q ENTENDER
                current_vlc_player.set_time(new_time_ms) #LINEA QUE LO ARRUINA (sin lo de arriba)
            
            # Actualizar la barra visualmente a la posición seleccionada por el usuario
            # Esto es importante si el set_time de VLC no es instantáneo o si hay alguna latencia.
            if progress_scale:
                program_is_setting_scale = True
                progress_scale.set(int(new_time_ms / 1000)) # Actualizar con el valor en segundos
                program_is_setting_scale = False
            
            if was_playing_before_seek:
                # Reiniciar las actualizaciones después de un breve retraso.
                ventana_principal.after(250, update_progress_bar) 
        else:
            print("No se puede buscar: duración de la canción desconocida o cero.")
    else:
        print("No se puede buscar: no hay reproductor o no está en estado válido.")

    #Para la cola de reproducción

#---------------------------------------------------------------------------------------------------------------




def destruir_etiqueta_si_existe(etiqueta):
    if etiqueta.winfo_exists():
        etiqueta.destroy()

def cerrar_ventana(ventana):
    ventana.destroy()

def alternar_visibilidad_contrasena(entrada):
    if entrada.cget("show") == "":
        entrada.config(show="*") 
    else:
        entrada.config(show="") 


def administrar_audio(archivo_audio,etiqueta_cancion_actual):
    global current_vlc_player
    global progress_scale
    global update_progress_job
    global imagen_label
    imagen_actualizada=archivo_audio.replace(".mp3", ".jpg")
    actualizar_imagen(imagen_actualizada)
    if not os.path.exists(archivo_audio):
        print(f"ERROR: El archivo '{archivo_audio}' no existe para reproducir.") # Added quotes
        return 

    nombre_cancion = os.path.basename(archivo_audio)
    nombre_cancion = nombre_cancion.replace(".mp3", "")
    etiqueta_cancion_actual.config(text=nombre_cancion)
    
    if current_vlc_player:
        current_vlc_player.stop()
        current_vlc_player.release() # Release previous current_vlc_player's resources
        current_vlc_player = None
        cancel_update_progress()

    current_vlc_player = vlc.MediaPlayer(archivo_audio)
    current_vlc_player.play()

    ventana_principal.after(100, lambda: setup_progress_bar(current_vlc_player))


def solo_escuchar(archivo_audio,nombre_cancion,etiqueta_cancion_actual): # Se eliminó el parámetro 'duracion'
    global current_vlc_player
    global imagen_label

    if os.path.exists(archivo_audio):
        etiqueta_cancion_actual.config(text=nombre_cancion)
        def check_and_delete():
            global current_vlc_player
            # Asegúrate de que el reproductor exista y esté en un estado válido para verificar
            if current_vlc_player and current_vlc_player.get_state() == vlc.State.Ended:
                print(f"DEBUG: Reproducción de '{archivo_audio}' finalizada (vlc.State.Ended). Procediendo a eliminar.")
                # Asegura que el reproductor esté detenido y liberado antes de intentar eliminar el archivo
                current_vlc_player.stop()
                current_vlc_player.release()
                current_vlc_player = None # Limpia la referencia
                
                # Intenta eliminar el archivo
                try:
                    ventana_principal.after(1000, lambda a=archivo_audio: os.remove(a))
                    print(f"DEBUG: Archivo '{archivo_audio}' eliminado exitosamente.")
                    imagen_label.config(image='') # Quita la imagen
                    imagen_label.image = None
                    archivo_imagen = ruta_main+'/'+'.jpg'
                    os.remove(archivo_imagen)
                except PermissionError as e:
                    print(f"ERROR: No se pudo eliminar el archivo '{archivo_audio}'. Aún en uso: {e}")
                    # Considera un mensaje de error más visible para el usuario si esto persiste
                except Exception as e:
                    print(f"ERROR inesperado al eliminar '{archivo_audio}': {e}")
                return # Detiene el chequeo recurrente una vez eliminado
            
            # Si no ha terminado, programa la próxima verificación en 1 segundo
            if current_vlc_player: # Continúa chequeando solo si el reproductor aún existe
                ventana_principal.after(1000, check_and_delete) # Vuelve a llamar en 1 segundo
            else:
                print("DEBUG: Reproductor VLC ya no existe, deteniendo chequeo de eliminación.")


        # Inicia el chequeo periódico
        ventana_principal.after(1000, check_and_delete) # Inicia el primer chequeo después de 1 segundo
        print(f"DEBUG: Archivo '{archivo_audio}' se reproducirá y se eliminará al finalizar.")
    else:
        print(f"ADVERTENCIA: No se encontró '{archivo_audio}' para reproducir o eliminar.")



def alt_pausa():
    global current_vlc_player
    global update_progress_job # Para manejar la actualización
    try:
        if current_vlc_player.get_state() == vlc.State.Playing:
            current_vlc_player.pause()
            cancel_update_progress()
        elif current_vlc_player.get_state() == vlc.State.Paused:
            current_vlc_player.play()
            update_progress_bar()
    except AttributeError:
        print("No se esta reproduciendo")
    return


try:
    icono_png = tk.PhotoImage(file="icono.png")
    ventana_principal.iconphoto(True, icono_png)
except tk.TclError:
    print("Error con el archivo del icono. Asegúrate de que 'icono.png' existe en la misma carpeta.")


etiqueta_p = tk.Label (ventana_principal,
                            text="Inicio de sesión",
                            foreground="#FFFFFF",
                            bg="#014942",
                            width="14",
                            height="1",
                            font="Corbel 67",
                            relief="groove"
                            )

etiqueta1 = tk.Label (ventana_principal,
                            text="Nombre de usuario:",
                            foreground="#FFFFFF",
                            bg="#014942",
                            font="Corbel 18",
                            )

entrada1 = tk.Entry (ventana_principal,
                            foreground="#FFFFFF",
                            bg="#01413A",
                            width="30",
                            font="Corbel 20",
                            )


etiqueta2 = tk.Label (ventana_principal,
                            text="Contraseña:",
                            foreground="#FFFFFF",
                            bg="#014942",
                            font="Corbel 18",
                            )

entrada2 = tk.Entry (ventana_principal,
                            foreground="#FFFFFF",
                            bg="#01413A",
                            width="30",
                            font="Corbel 20",
                            show="*"
                            )

boton_ver = tk.Button(ventana_principal,
                            text="Ver",
                            command=lambda: alternar_visibilidad_contrasena(entrada2),
                            bg="#014942",
                            foreground="#FFFFFF",
                            font="Corbel 10",
                            )


boton1 = tk.Button (ventana_principal,
                            text="Iniciar sesión",
                            foreground="#FFFFFF",
                            bg="#007C70",
                            width="25",
                            height="1",
                            font="Corbel 30",
                            justify="center",
                            activebackground="#06D6C1",
                            activeforeground="#FFFFFF",
                            command=inicio_sesion
                            )

boton2 = tk.Button (ventana_principal,text="¿No tiene cuenta? Registrese",
                            foreground="#FFFFFF",
                            bg="#014942",
                            width="25",
                            height="1",
                            font="Corbel 10",
                            justify="center",
                            activebackground="#059E8F",
                            activeforeground="#FFFFFF",
                            command=crear_cuenta
                            )


etiqueta3 = tk.Label (ventana_principal,
                            text="Ingrese datos validos",
                            foreground="#FF0000",
                            bg="#014942",
                            font="Corbel 18",
                            )

etiqueta_p.pack(pady=60,padx=40)
etiqueta1.pack()
entrada1.pack()
etiqueta2.pack()
entrada2.pack()
boton_ver.pack()
boton1.pack(pady=20)
boton2.pack(pady=20)

ventana_principal.mainloop()
